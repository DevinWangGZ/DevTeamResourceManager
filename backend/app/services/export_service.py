"""数据导出服务"""
import io
import re
from urllib.request import urlopen
from urllib.parse import urlparse, unquote
from datetime import datetime, date
from typing import Optional
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from PIL import Image as PILImage

from app.services.task_service import TaskService
from app.schemas.task import TaskFilterParams
from app.models.task import TaskStatus
from app.models.workload_statistic import WorkloadStatistic
from app.models.user import User
from app.models.project import Project
from app.utils.paths import get_uploads_dir


_MARKDOWN_IMAGE = re.compile(r"!\[[^\]]*\]\(([^)]+)\)")
_IMAGE_FETCH_TIMEOUT_SECONDS = 8
_MAX_IMAGES_PER_TASK = 3


class ExportService:
    """数据导出服务类"""

    @staticmethod
    def _description_for_excel(description: Optional[str]) -> tuple[str, list[str]]:
        """
        「问题内容」在库中多为 Markdown：图片为 ![](url) / ![alt](url)。
        这里会把图片 URL 提取出来，并将正文中的 Markdown 图片语法移除，
        图片本体在导出时尝试嵌入到“问题图片”列。
        """
        if not description:
            return "", []
        text = description.strip()
        if not text:
            return "", []
        image_urls = [url.strip() for url in _MARKDOWN_IMAGE.findall(text) if url.strip()]
        cleaned = _MARKDOWN_IMAGE.sub("", text)
        # 清理图片语法留下的空白行，提升可读性
        cleaned = re.sub(r"\n\s*\n+", "\n\n", cleaned).strip()
        return cleaned, image_urls

    @staticmethod
    def _fetch_image_bytes(url: str) -> Optional[io.BytesIO]:
        """下载图片并返回内存流；失败时返回 None，避免导出整体失败。"""
        parsed = urlparse(url)
        uploads_dir = get_uploads_dir().resolve()

        # 站内图片优先直接读本地文件，避免经公网回环导致慢/超时
        if parsed.path and parsed.path.startswith("/uploads/"):
            try:
                relative = parsed.path[len("/uploads/"):].lstrip("/")
                local_path = (uploads_dir / unquote(relative)).resolve()
                if local_path.is_file() and uploads_dir in local_path.parents:
                    data = local_path.read_bytes()
                    if data:
                        return io.BytesIO(data)
            except Exception:
                pass

        try:
            with urlopen(url, timeout=_IMAGE_FETCH_TIMEOUT_SECONDS) as resp:
                data = resp.read()
            if not data:
                return None
            return io.BytesIO(data)
        except Exception:
            return None

    @staticmethod
    def _scale_image(img: XLImage, max_width: int = 220, max_height: int = 140) -> None:
        """按比例缩放图片，限制在单元格可视区域。"""
        width = float(img.width or 0)
        height = float(img.height or 0)
        if width <= 0 or height <= 0:
            return
        scale = min(max_width / width, max_height / height, 1.0)
        img.width = int(width * scale)
        img.height = int(height * scale)

    @staticmethod
    def _add_images_to_same_cell(
        ws,
        row_idx: int,
        col_idx: int,
        image_urls: list[str],
        image_stream_refs: list[io.BytesIO],
    ) -> None:
        """
        在同一单元格内展示多图：先合成为一张竖向拼接图，再插入单元格。
        这样兼容性更高（WPS/Excel 都稳定），也能满足“同单元格多图”诉求。
        """
        if not image_urls:
            return

        images: list[PILImage.Image] = []
        gap_px = 8
        max_width_px = 210
        per_image_max_height_px = 120
        for url in image_urls[:_MAX_IMAGES_PER_TASK]:
            image_buffer = ExportService._fetch_image_bytes(url)
            if not image_buffer:
                continue
            try:
                pil = PILImage.open(image_buffer)
                pil.load()
                pil = pil.convert("RGB")
                width, height = pil.size
                if width <= 0 or height <= 0:
                    continue
                scale = min(max_width_px / width, per_image_max_height_px / height, 1.0)
                new_size = (max(1, int(width * scale)), max(1, int(height * scale)))
                if new_size != pil.size:
                    pil = pil.resize(new_size, PILImage.Resampling.LANCZOS)
                images.append(pil)
            except Exception:
                continue

        if not images:
            return

        total_height = sum(img.height for img in images) + gap_px * (len(images) - 1)
        max_width = max(img.width for img in images)
        canvas = PILImage.new("RGB", (max_width, total_height), color=(255, 255, 255))

        y = 0
        for img in images:
            x = (max_width - img.width) // 2
            canvas.paste(img, (x, y))
            y += img.height + gap_px

        merged_buffer = io.BytesIO()
        canvas.save(merged_buffer, format="PNG")
        merged_buffer.seek(0)
        image_stream_refs.append(merged_buffer)

        excel_image = XLImage(merged_buffer)
        ws.add_image(excel_image, f"{get_column_letter(col_idx)}{row_idx}")
        ws.row_dimensions[row_idx].height = max(
            ws.row_dimensions[row_idx].height or 15,
            total_height * 0.75 + 6,
        )

    @staticmethod
    def _create_header_style():
        """创建表头样式"""
        return {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    @staticmethod
    def _apply_header_style(ws, row, max_col):
        """应用表头样式"""
        header_style = ExportService._create_header_style()
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            for attr, value in header_style.items():
                setattr(cell, attr, value)

    @staticmethod
    def export_workload_statistics(
        db: Session,
        user_id: Optional[int] = None,
        project_id: Optional[int] = None,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> io.BytesIO:
        """导出工作量统计数据到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "工作量统计"

        # 表头
        headers = [
            "统计周期", "用户", "项目", "任务数", "投入人天", "完成率", "创建时间"
        ]
        ws.append(headers)
        ExportService._apply_header_style(ws, 1, len(headers))

        # 查询数据
        query = db.query(WorkloadStatistic)
        if user_id:
            query = query.filter(WorkloadStatistic.user_id == user_id)
        if project_id:
            query = query.filter(WorkloadStatistic.project_id == project_id)
        if period_start:
            query = query.filter(WorkloadStatistic.period_start >= period_start)
        if period_end:
            query = query.filter(WorkloadStatistic.period_end <= period_end)

        statistics = query.order_by(WorkloadStatistic.period_start.desc()).all()

        # 填充数据
        for stat in statistics:
            user = db.query(User).filter(User.id == stat.user_id).first()
            project = db.query(Project).filter(Project.id == stat.project_id).first() if stat.project_id else None

            completion_rate = "0%"
            if stat.total_tasks > 0:
                completion_rate = f"{(stat.completed_tasks / stat.total_tasks * 100):.1f}%"

            ws.append([
                f"{stat.period_start} 至 {stat.period_end}",
                user.full_name or user.username if user else "未知",
                project.name if project else "未分配项目",
                stat.total_tasks,
                float(stat.total_man_days),
                completion_rate,
                stat.created_at.strftime("%Y-%m-%d %H:%M:%S") if stat.created_at else ""
            ])

        # 调整列宽
        column_widths = [20, 15, 20, 10, 12, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_tasks(
        db: Session,
        filters: TaskFilterParams,
        current_user_id: int,
        current_user_role: str,
        embed_images: bool = False,
    ) -> io.BytesIO:
        """导出任务数据到Excel（筛选条件与任务列表 API 一致，最多 10000 条）。"""
        dumped = filters.model_dump()
        dumped["page"] = 1
        dumped["page_size"] = 10000
        export_filters = TaskFilterParams.model_construct(**dumped)
        tasks, _ = TaskService.get_tasks(
            db,
            export_filters,
            current_user_id=current_user_id,
            current_user_role=current_user_role,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "任务列表"

        # 表头（「问题内容」对应模型 description，与任务描述一致）
        headers = [
            "任务ID", "标题", "问题内容", "问题图片", "项目", "状态", "创建人", "负责人", "拟投入人天", "实际投入人天",
            "创建时间", "更新时间"
        ]
        ws.append(headers)
        ExportService._apply_header_style(ws, 1, len(headers))
        image_stream_refs: list[io.BytesIO] = []

        # 填充数据
        for row_idx, task in enumerate(tasks, start=2):
            # 使用已加载的关联对象
            creator = task.creator
            assignee = task.assignee
            project = task.project

            status_text = {
                TaskStatus.DRAFT.value: "草稿",
                TaskStatus.PUBLISHED.value: "已发布",
                TaskStatus.CLAIMED.value: "已认领",
                TaskStatus.PENDING_EVAL.value: "待评估",
                TaskStatus.IN_PROGRESS.value: "进行中",
                TaskStatus.SUBMITTED.value: "已提交",
                TaskStatus.CONFIRMED.value: "已确认",
                TaskStatus.ARCHIVED.value: "已归档",
            }.get(task.status, task.status)

            desc, image_urls = ExportService._description_for_excel(task.description)
            image_cell_value = "\n".join(image_urls) if image_urls else ""
            ws.append([
                task.id,
                task.title,
                desc,
                image_cell_value,
                project.name if project else "未分配项目",
                status_text,
                creator.full_name or creator.username if creator else "未知",
                assignee.full_name or assignee.username if assignee else "未分配",
                float(task.estimated_man_days) if task.estimated_man_days else 0,
                float(task.actual_man_days) if task.actual_man_days else 0,
                task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else "",
                task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else ""
            ])

            # 默认允许正文换行显示
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")

            # 开启 embed_images 时才嵌图；默认关闭，仅保留 URL 文本，导出速度更可控。
            if embed_images:
                ExportService._add_images_to_same_cell(
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=4,
                    image_urls=image_urls,
                    image_stream_refs=image_stream_refs,
                )

        # 调整列宽
        column_widths = [10, 28, 40, 18, 20, 12, 15, 15, 12, 12, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_performance_data(
        db: Session,
        user_id: Optional[int] = None,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> io.BytesIO:
        """导出绩效数据到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "绩效数据"

        # 表头
        headers = [
            "用户", "序列等级", "统计周期", "完成任务数", "总任务数", "完成率",
            "总投入人天", "平均投入人天/任务", "创建时间"
        ]
        ws.append(headers)
        ExportService._apply_header_style(ws, 1, len(headers))

        # 查询数据
        query = db.query(WorkloadStatistic)
        if user_id:
            query = query.filter(WorkloadStatistic.user_id == user_id)
        if period_start:
            query = query.filter(WorkloadStatistic.period_start >= period_start)
        if period_end:
            query = query.filter(WorkloadStatistic.period_end <= period_end)

        statistics = query.order_by(WorkloadStatistic.period_start.desc()).all()

        # 填充数据
        for stat in statistics:
            user = db.query(User).filter(User.id == stat.user_id).first()
            if not user:
                continue

            # 获取序列等级
            from app.models.user_sequence import UserSequence
            sequence = db.query(UserSequence).filter(
                UserSequence.user_id == user.id
            ).order_by(UserSequence.created_at.desc()).first()
            sequence_level = sequence.level if sequence else "未设置"

            completion_rate = "0%"
            if stat.total_tasks > 0:
                completion_rate = f"{(stat.completed_tasks / stat.total_tasks * 100):.1f}%"

            avg_man_days = "0"
            if stat.completed_tasks > 0:
                avg_man_days = f"{(float(stat.total_man_days) / stat.completed_tasks):.2f}"

            ws.append([
                user.full_name or user.username,
                sequence_level,
                f"{stat.period_start} 至 {stat.period_end}",
                stat.completed_tasks,
                stat.total_tasks,
                completion_rate,
                float(stat.total_man_days),
                avg_man_days,
                stat.created_at.strftime("%Y-%m-%d %H:%M:%S") if stat.created_at else ""
            ])

        # 调整列宽
        column_widths = [15, 12, 25, 12, 12, 10, 12, 18, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
